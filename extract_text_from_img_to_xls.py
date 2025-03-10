import os
import pytesseract
from PIL import Image
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

def create_excel_file():
    wb = Workbook()
    ws = wb.active
    
    headers = [
        'Document ID', 'Page', 'Position', 'Filename',
        'Text Content', 'Date', 'Characters', 'Words', 'Path', 'Source Folder'
    ]
    
    # Style d'en-tête
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    
    # Ajuster largeur des colonnes
    ws.column_dimensions['A'].width = 20  # Document ID
    ws.column_dimensions['B'].width = 8   # Page
    ws.column_dimensions['C'].width = 10  # Position
    ws.column_dimensions['D'].width = 20  # Filename
    ws.column_dimensions['E'].width = 50  # Text Content
    ws.column_dimensions['F'].width = 20  # Date
    ws.column_dimensions['G'].width = 12  # Characters
    ws.column_dimensions['H'].width = 8   # Words
    ws.column_dimensions['I'].width = 50  # Path
    ws.column_dimensions['J'].width = 15  # Source Folder
    
    return wb, ws

def process_image(file_path, filename, td_folder, source_folder, ws, row):
    try:
        parts = filename.replace('.jpg', '').split('_')
        page_number = parts[1] if len(parts) > 1 else ''
        page_position = parts[2] if len(parts) > 2 else ''
        
        img = Image.open(file_path)
        text = pytesseract.image_to_string(img, lang='fra')
        text_stripped = text.strip()
        
        # Remplir la ligne
        ws.cell(row=row, column=1, value=td_folder)
        ws.cell(row=row, column=2, value=page_number)
        ws.cell(row=row, column=3, value=page_position)
        ws.cell(row=row, column=4, value=filename)
        ws.cell(row=row, column=5, value=text_stripped)
        ws.cell(row=row, column=6, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row, column=7, value=len(text_stripped))
        ws.cell(row=row, column=8, value=len(text_stripped.split()))
        ws.cell(row=row, column=9, value=file_path)
        ws.cell(row=row, column=10, value=source_folder)
        
        print(f"Processed: {filename} from {source_folder}")
        return True
        
    except Exception as e:
        print(f"Error processing {file_path}: {e}")
        return False

def extract_text_from_images():
    excel_file = 'extracted_text.xlsx'
    base_dir = 'output_images'
    
    wb, ws = create_excel_file()
    row = 2  # Start after header
    
    for td_folder in os.listdir(base_dir):
        td_path = os.path.join(base_dir, td_folder)
        
        # Traiter les images dans doubles_images_pages/split_images
        doubles_dir = os.path.join(td_path, 'doubles_images_pages', 'split_images')
        if os.path.exists(doubles_dir):
            for filename in os.listdir(doubles_dir):
                if filename.lower().endswith(('.jpg', '.jpeg', '.png')):
                    file_path = os.path.join(doubles_dir, filename)
                    if process_image(file_path, filename, td_folder, 'doubles_images_pages', ws, row):
                        row += 1
                    
                    # Sauvegarder régulièrement
                    if row % 10 == 0:
                        wb.save(excel_file)
        
        # Traiter les images dans une_image_page
        une_image_dir = os.path.join(td_path, 'une_image_page')
        if os.path.exists(une_image_dir):
            for filename in os.listdir(une_image_dir):
                if filename.lower().endswith(('.jpg', '.jpeg', '.png')):
                    file_path = os.path.join(une_image_dir, filename)
                    if process_image(file_path, filename, td_folder, 'une_image_page', ws, row):
                        row += 1
                    
                    # Sauvegarder régulièrement
                    if row % 10 == 0:
                        wb.save(excel_file)
    
    wb.save(excel_file)
    print(f"Excel file saved: {excel_file}")

if __name__ == "__main__":
    extract_text_from_images()